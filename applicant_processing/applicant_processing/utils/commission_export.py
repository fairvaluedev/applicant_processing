# Copyright (c) 2026, Admin and contributors
# For license information, please see license.txt

import io
import json
import csv
import frappe
from frappe.utils import getdate, today, formatdate, flt


# ─────────────────────────────────────────────────────────────────────────────
# Internal Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _commission_columns_present():
    """
    Check once whether the commission columns exist in the Applicant table.
    Returns a dict of booleans keyed by column name.
    """
    cols = ["commission_status", "commission_amount", "commission_paid_date", "commission_batch_ref"]
    result = {}
    for col in cols:
        try:
            result[col] = bool(frappe.db.has_column("Applicant", col))
        except Exception:
            result[col] = False
    return result


def _normalize_limit(limit):
    """
    Parse the limit parameter into an integer or None (meaning ALL).
    Returns: (int | None, str label)
    """
    if not limit or str(limit).lower() in ("all", "0", "none", ""):
        return None, "All"
    try:
        n = int(limit)
        return (n if n > 0 else None), str(n)
    except (ValueError, TypeError):
        return None, "All"


def _parse_applicant_ids(applicant_ids):
    """Deserialize applicant_ids from JSON string, comma-list, or list."""
    if not applicant_ids:
        return []
    if isinstance(applicant_ids, list):
        return [a for a in applicant_ids if a]
    if isinstance(applicant_ids, str):
        try:
            parsed = json.loads(applicant_ids)
            if isinstance(parsed, list):
                return [a for a in parsed if a]
        except (json.JSONDecodeError, ValueError):
            pass
        return [a.strip() for a in applicant_ids.split(",") if a.strip()]
    return []


def _get_contractor(contractor):
    """Fetch the Contractor doc with a clean error on miss."""
    if not contractor:
        frappe.throw(frappe._("Contractor is required."))
    if not frappe.db.exists("Contractor", contractor):
        frappe.throw(frappe._("Partner agency '{0}' was not found.").format(contractor), frappe.DoesNotExistError)
    return frappe.get_doc("Contractor", contractor)


# ─────────────────────────────────────────────────────────────────────────────
# Core Data Fetch
# ─────────────────────────────────────────────────────────────────────────────

def get_unpaid_commission_data(contractor, limit=30, from_date=None, to_date=None, applicant_ids=None):
    """
    Fetches departed applicants for a specific foreign recruitment agency whose
    commission has not been settled.

    Returns: (summary dict, list of candidate dicts)
    """
    contractor_doc = _get_contractor(contractor)
    default_rate = flt(contractor_doc.default_commission_amount) or 0.0
    currency = contractor_doc.default_commission_currency or "SAR"

    cols = _commission_columns_present()
    has_status = cols["commission_status"]
    has_amount = cols["commission_amount"]

    # Build SQL fragments that gracefully degrade if columns don't exist yet
    status_select = "app.commission_status" if has_status else "'Unpaid'"
    amount_select = "app.commission_amount" if has_amount else "NULL"

    status_filter = (
        "(app.commission_status IS NULL OR app.commission_status = '' OR app.commission_status = 'Unpaid')"
        if has_status else "1=1"
    )

    conditions = [
        "(app.locked_contractor = %(contractor)s OR dos.contractor_name = %(contractor)s OR dsr.contractor_name = %(contractor)s)",
        "(app.applicant_state = 'Departed' OR dep.status = 'Departed' OR dsr.departure_status = 'Departed')",
        status_filter,
    ]
    params = {"contractor": contractor}

    ids = _parse_applicant_ids(applicant_ids)
    if ids:
        conditions.append("app.name IN %(applicant_ids)s")
        params["applicant_ids"] = tuple(ids)

    if from_date:
        conditions.append("DATE(COALESCE(dep.departure_time, dep.modified, app.modified)) >= %(from_date)s")
        params["from_date"] = str(getdate(from_date))

    if to_date:
        conditions.append("DATE(COALESCE(dep.departure_time, dep.modified, app.modified)) <= %(to_date)s")
        params["to_date"] = str(getdate(to_date))

    where = " AND ".join(conditions)
    limit_int, limit_label = _normalize_limit(limit)
    limit_clause = f"LIMIT {limit_int}" if limit_int else ""

    sql = f"""
        SELECT
            app.name,
            app.full_name,
            app.first_name,
            app.last_name,
            app.passport_number,
            app.destination_country,
            app.job_applied,
            app.phone_number,
            app.applicant_state,
            {status_select} AS commission_status,
            {amount_select} AS commission_amount,
            COALESCE(dep.departure_time, dep.modified, app.modified) AS departure_date,
            dos.name AS dossier,
            dos.sponsor_name,
            dos.visa_number
        FROM `tabApplicant` app
        LEFT JOIN `tabApplicant Dossier` dos ON dos.applicant = app.name
        LEFT JOIN `tabDSR` dsr ON dsr.applicant_dossier = dos.name
        LEFT JOIN `tabDSR Departure` dep ON dep.dsr = dsr.name
        WHERE {where}
        GROUP BY app.name
        ORDER BY departure_date DESC, app.creation DESC
        {limit_clause}
    """

    rows = frappe.db.sql(sql, params, as_dict=True)

    candidates = []
    total_amount = 0.0

    for idx, row in enumerate(rows, 1):
        rate = flt(row.get("commission_amount")) or default_rate
        total_amount += rate

        dep_dt = row.get("departure_date")
        dep_str = ""
        if dep_dt:
            try:
                dep_str = formatdate(dep_dt, "dd-MMM-yyyy")
            except Exception:
                dep_str = str(dep_dt)[:10]

        full_name = (
            row.get("full_name")
            or f"{row.get('first_name', '')} {row.get('last_name', '')}".strip()
            or "—"
        )

        candidates.append({
            "idx": idx,
            "name": row.get("name"),
            "full_name": full_name,
            "passport_number": row.get("passport_number") or "—",
            "destination_country": row.get("destination_country") or contractor_doc.country or "—",
            "job_applied": row.get("job_applied") or "—",
            "phone_number": row.get("phone_number") or "—",
            "sponsor_name": row.get("sponsor_name") or "—",
            "visa_number": row.get("visa_number") or "—",
            "departure_date": dep_str,
            "commission_rate": rate,
            "commission_currency": currency,
            "commission_status": "Unpaid",
        })

    batch_label = (
        f"Last {len(candidates)} of {limit_label}"
        if limit_int and len(candidates) == limit_int
        else f"All ({len(candidates)})"
    )

    summary = {
        "contractor": contractor,
        "company_name": contractor_doc.company_name,
        "country": contractor_doc.country or "—",
        "contact_person": contractor_doc.contact_person or "—",
        "phone": contractor_doc.phone or contractor_doc.whatsapp or "—",
        "email": contractor_doc.email or "—",
        "default_rate": default_rate,
        "currency": currency,
        "total_count": len(candidates),
        "total_amount": total_amount,
        "generated_on": formatdate(today(), "dd-MMM-yyyy"),
        "batch_label": batch_label,
    }

    return summary, candidates


# ─────────────────────────────────────────────────────────────────────────────
# Excel Builder
# ─────────────────────────────────────────────────────────────────────────────

def build_commission_excel(summary, candidates):
    """
    Generates a styled Excel workbook (.xlsx) using openpyxl.
    Falls back to UTF-8 CSV if openpyxl is not available.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commission Statement"

        # ── Colour tokens ────────────────────────────────────────────────────
        C_NAVY   = "1E293B"
        C_TEAL   = "0F766E"
        C_ACCENT = "EFF6FF"
        C_ALT    = "F1F5F9"
        C_WHITE  = "FFFFFF"
        C_TOTAL  = "ECFDF5"

        def fill(hex_color):
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        def border(color="CBD5E1"):
            s = Side(style="thin", color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        def font(bold=False, color="0F172A", size=10, name="Calibri"):
            return Font(name=name, size=size, bold=bold, color=color)

        def align(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        # ── Title row (merged A1:K2) ─────────────────────────────────────────
        ws.merge_cells("A1:K2")
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 22
        t = ws["A1"]
        t.value = f"COMMISSION BILLING STATEMENT  —  {summary['company_name'].upper()}"
        t.font = Font(name="Calibri", size=14, bold=True, color=C_WHITE)
        t.fill = fill(C_NAVY)
        t.alignment = align("center")

        # ── Metadata rows (4–6) ──────────────────────────────────────────────
        meta = [
            ("A4", "Partner Agency:",       "B4", summary["company_name"]),
            ("D4", "Operating Country:",    "E4", summary["country"]),
            ("G4", "Statement Date:",       "H4", summary["generated_on"]),
            ("A5", "Contact Person:",       "B5", summary["contact_person"]),
            ("D5", "Phone / WhatsApp:",     "E5", summary["phone"]),
            ("G5", "Scope:",                "H5", summary["batch_label"]),
            ("A6", "Rate / Candidate:",     "B6", f"{summary['default_rate']:,.2f} {summary['currency']}"),
            ("D6", "Unpaid Candidates:",    "E6", summary["total_count"]),
            ("G6", "Total Outstanding:",    "H6", f"{summary['total_amount']:,.2f} {summary['currency']}"),
        ]
        for label_cell, label_val, val_cell, val_val in meta:
            ws[label_cell].value = label_val
            ws[label_cell].font = font(bold=True, color=C_TEAL)
            ws[label_cell].fill = fill(C_ACCENT)
            ws[val_cell].value = val_val
            ws[val_cell].font = font()

        # ── Column headers (row 8) ────────────────────────────────────────────
        headers = [
            "#", "Applicant ID", "Full Name", "Passport No",
            "Destination", "Job Position", "Departure Date",
            "Sponsor / Employer", "Visa No", f"Commission ({summary['currency']})", "Status"
        ]
        COL_WIDTHS = [6, 14, 22, 14, 14, 18, 14, 22, 14, 18, 12]
        for i, (h, w) in enumerate(zip(headers, COL_WIDTHS), 1):
            c = ws.cell(row=8, column=i, value=h)
            c.font = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
            c.fill = fill(C_TEAL)
            c.alignment = align("center")
            c.border = border()
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[8].height = 18

        # ── Data rows (row 9+) ────────────────────────────────────────────────
        for row_num, cand in enumerate(candidates, 9):
            row_fill = fill(C_ALT) if row_num % 2 == 0 else fill(C_WHITE)
            row_vals = [
                cand["idx"], cand["name"], cand["full_name"], cand["passport_number"],
                cand["destination_country"], cand["job_applied"], cand["departure_date"],
                cand["sponsor_name"], cand["visa_number"], cand["commission_rate"], cand["commission_status"]
            ]
            for col_num, val in enumerate(row_vals, 1):
                c = ws.cell(row=row_num, column=col_num, value=val)
                c.font = font()
                c.fill = row_fill
                c.border = border()
                if col_num in (1, 4, 7, 11):
                    c.alignment = align("center")
                elif col_num == 10:
                    c.alignment = align("right")
                    c.number_format = "#,##0.00"
                else:
                    c.alignment = align("left")

        # ── Total footer row ─────────────────────────────────────────────────
        total_row = 9 + len(candidates)
        ws.merge_cells(
            start_row=total_row, start_column=1,
            end_row=total_row, end_column=9
        )
        lbl = ws.cell(row=total_row, column=1)
        lbl.value = f"TOTAL COMMISSION PAYABLE  ({summary['total_count']} candidates)"
        lbl.font = Font(name="Calibri", size=11, bold=True, color=C_TEAL)
        lbl.fill = fill(C_TOTAL)
        lbl.alignment = align("right")

        amt = ws.cell(row=total_row, column=10)
        amt.value = summary["total_amount"]
        amt.font = Font(name="Calibri", size=12, bold=True, color=C_TEAL)
        amt.fill = fill(C_TOTAL)
        amt.alignment = align("right")
        amt.number_format = f'#,##0.00 "{summary["currency"]}"'

        st = ws.cell(row=total_row, column=11)
        st.value = "UNPAID"
        st.font = Font(name="Calibri", size=10, bold=True, color="B91C1C")
        st.fill = fill(C_TOTAL)
        st.alignment = align("center")

        for col in range(1, 12):
            ws.cell(row=total_row, column=col).border = border()

        # ── Freeze pane & sheet settings ─────────────────────────────────────
        ws.freeze_panes = "A9"
        ws.auto_filter.ref = f"A8:K{total_row - 1}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"

    except Exception as exc:
        frappe.log_error(title="Commission Excel build failed; falling back to CSV", message=str(exc))

    # ── CSV fallback ─────────────────────────────────────────────────────────
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f"COMMISSION BILLING STATEMENT — {summary['company_name']}"])
    w.writerow(["Date", summary["generated_on"], "Country", summary["country"],
                "Rate", f"{summary['default_rate']} {summary['currency']}"])
    w.writerow([])
    w.writerow(["#", "Applicant ID", "Full Name", "Passport No", "Destination",
                "Job", "Departure", "Sponsor", "Visa No",
                f"Commission ({summary['currency']})", "Status"])
    for c in candidates:
        w.writerow([c["idx"], c["name"], c["full_name"], c["passport_number"],
                    c["destination_country"], c["job_applied"], c["departure_date"],
                    c["sponsor_name"], c["visa_number"], c["commission_rate"], c["commission_status"]])
    w.writerow([])
    w.writerow(["TOTAL", "", "", "", "", "", "", "", "", summary["total_amount"], "UNPAID"])
    return buf.getvalue().encode("utf-8-sig"), "text/csv", "csv"


# ─────────────────────────────────────────────────────────────────────────────
# PDF Builder
# ─────────────────────────────────────────────────────────────────────────────

def build_commission_pdf(summary, candidates):
    """
    Renders an A4 PDF billing statement via wkhtmltopdf.
    """
    from frappe.utils.pdf import get_pdf

    rows_html = "".join(
        f"""<tr>
            <td class="tc b">{c['idx']}</td>
            <td class="b">{frappe.utils.escape_html(c['name'])}</td>
            <td>{frappe.utils.escape_html(c['full_name'])}</td>
            <td class="mono">{frappe.utils.escape_html(c['passport_number'])}</td>
            <td>{frappe.utils.escape_html(c['job_applied'])}</td>
            <td class="tc">{frappe.utils.escape_html(c['departure_date'])}</td>
            <td>{frappe.utils.escape_html(c['sponsor_name'])}</td>
            <td class="tr b">{c['commission_rate']:,.2f} {frappe.utils.escape_html(c['commission_currency'])}</td>
        </tr>"""
        for c in candidates
    )

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  @page {{ size: A4; margin: 14mm 16mm 16mm 16mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: "Helvetica Neue", Helvetica, Arial, sans-serif; font-size: 10px; color: #0f172a; margin: 0; }}
  .page-header {{ border-bottom: 2px solid #0f766e; padding-bottom: 10px; margin-bottom: 14px; overflow: hidden; }}
  .page-header .left {{ float: left; }}
  .page-header .right {{ float: right; text-align: right; }}
  .doc-title {{ font-size: 16px; font-weight: 800; color: #0f766e; letter-spacing: -0.02em; margin: 0 0 2px; }}
  .doc-sub {{ font-size: 10px; color: #64748b; margin: 0; }}
  .badge-inv {{ background: #0f766e; color: #fff; padding: 5px 11px; border-radius: 5px; font-size: 11px; font-weight: 700; display: inline-block; }}
  .meta-grid {{ display: table; width: 100%; border: 1px solid #e2e8f0; border-radius: 6px; margin-bottom: 14px; background: #f8fafc; }}
  .meta-col {{ display: table-cell; padding: 9px 13px; border-right: 1px solid #e2e8f0; vertical-align: top; width: 25%; }}
  .meta-col:last-child {{ border-right: none; background: #ecfdf5; }}
  .meta-label {{ font-size: 8px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.06em; color: #64748b; margin-bottom: 3px; }}
  .meta-val {{ font-size: 12px; font-weight: 800; color: #0f172a; }}
  .meta-sub {{ font-size: 8px; color: #64748b; margin-top: 1px; }}
  .meta-total-val {{ font-size: 14px; font-weight: 800; color: #047857; }}
  table.dt {{ width: 100%; border-collapse: collapse; margin-bottom: 12px; }}
  table.dt thead tr {{ background: #1e293b; color: #fff; }}
  table.dt thead th {{ font-size: 8px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em; padding: 6px 5px; text-align: left; border: 1px solid #1e293b; }}
  table.dt tbody tr:nth-child(even) {{ background: #f8fafc; }}
  table.dt td {{ padding: 5px 5px; font-size: 9.5px; border: 1px solid #e2e8f0; }}
  table.dt tfoot td {{ background: #ecfdf5; font-weight: 800; font-size: 10px; padding: 7px 5px; border: 1px solid #86efac; color: #047857; }}
  .tc {{ text-align: center; }}
  .tr {{ text-align: right; }}
  .b {{ font-weight: 700; }}
  .mono {{ font-family: "Courier New", monospace; font-size: 9px; }}
  .notice {{ font-size: 8.5px; color: #64748b; line-height: 1.5; border-top: 1px dashed #cbd5e1; padding-top: 8px; margin-top: 8px; }}
  .sig-table {{ display: table; width: 100%; margin-top: 28px; }}
  .sig-cell {{ display: table-cell; width: 48%; border-top: 1px solid #94a3b8; padding-top: 5px; font-size: 9px; color: #475569; }}
</style>
</head>
<body>

<div class="page-header">
  <div class="left">
    <p class="doc-title">RECRUITMENT COMMISSION STATEMENT</p>
    <p class="doc-sub">Automated Billing &amp; Settlement Engine</p>
  </div>
  <div class="right">
    <div class="badge-inv">BILLING STATEMENT</div>
    <div style="font-size:9px;color:#64748b;margin-top:3px;">Date: {summary['generated_on']}</div>
  </div>
</div>

<div class="meta-grid">
  <div class="meta-col">
    <div class="meta-label">Partner Agency</div>
    <div class="meta-val">{frappe.utils.escape_html(summary['company_name'])}</div>
    <div class="meta-sub">{frappe.utils.escape_html(summary['country'])} &nbsp;|&nbsp; {frappe.utils.escape_html(summary['contact_person'])}</div>
  </div>
  <div class="meta-col">
    <div class="meta-label">Rate / Candidate</div>
    <div class="meta-val">{summary['default_rate']:,.2f} {summary['currency']}</div>
    <div class="meta-sub">Scope: {frappe.utils.escape_html(summary['batch_label'])}</div>
  </div>
  <div class="meta-col">
    <div class="meta-label">Departed (Unpaid)</div>
    <div class="meta-val">{summary['total_count']} Candidates</div>
    <div class="meta-sub" style="color:#b91c1c;font-weight:700;">Status: UNPAID</div>
  </div>
  <div class="meta-col">
    <div class="meta-label">Total Commission Due</div>
    <div class="meta-total-val">{summary['total_amount']:,.2f} {summary['currency']}</div>
    <div class="meta-sub" style="color:#047857;">Due upon receipt</div>
  </div>
</div>

<table class="dt">
  <thead>
    <tr>
      <th style="width:4%;" class="tc">#</th>
      <th style="width:12%;">Candidate ID</th>
      <th style="width:20%;">Full Name</th>
      <th style="width:12%;">Passport No</th>
      <th style="width:13%;">Job Position</th>
      <th style="width:10%;" class="tc">Departed</th>
      <th style="width:17%;">Sponsor / Employer</th>
      <th style="width:12%;" class="tr">Commission</th>
    </tr>
  </thead>
  <tbody>
    {rows_html}
  </tbody>
  <tfoot>
    <tr>
      <td colspan="7" class="tr">TOTAL OUTSTANDING ({summary['total_count']} candidates):</td>
      <td class="tr">{summary['total_amount']:,.2f} {summary['currency']}</td>
    </tr>
  </tfoot>
</table>

<div class="notice">
  <strong>Note:</strong> This statement covers workers who have departed under the
  <strong>{frappe.utils.escape_html(summary['company_name'])}</strong> placement quota.
  Commission is due at the agreed rate of
  <strong>{summary['default_rate']:,.2f} {summary['currency']}</strong> per deployed candidate.
  This document constitutes an official billing request.
</div>

<div class="sig-table">
  <div class="sig-cell"><strong>Prepared By:</strong> Finance &amp; Accounts — Applicant Processing System</div>
  <div class="sig-cell" style="text-align:right;">
    <strong>Received &amp; Acknowledged:</strong><br>
    {frappe.utils.escape_html(summary['company_name'])}
  </div>
</div>

</body>
</html>"""

    pdf_bytes = get_pdf(html, options={
        "page-size": "A4",
        "orientation": "Portrait",
        "margin-top": "14mm",
        "margin-bottom": "14mm",
        "margin-left": "16mm",
        "margin-right": "16mm",
        "encoding": "UTF-8",
    })
    return pdf_bytes, "application/pdf", "pdf"


# ─────────────────────────────────────────────────────────────────────────────
# Whitelisted API Endpoints
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_session_contractor(contractor=None):
    """
    Enforces multi-tenant security on commission endpoints.
    If user has Foreign Agency role, strictly resolves contractor to their own linked Contractor doc.
    """
    if not frappe.session.user or frappe.session.user == "Guest":
        frappe.throw("Authentication required. Please log in.", frappe.AuthenticationError)

    user_roles = frappe.get_roles(frappe.session.user)
    is_internal = any(r in user_roles for r in ("System Manager", "Administrator", "LMS Employee", "Accounts Manager"))

    if is_internal and contractor:
        return contractor

    contractor_name = frappe.db.get_value("Contractor", {"email": frappe.session.user, "active_status": 1}, "name")
    if not contractor_name and hasattr(frappe.db, "has_column") and frappe.db.has_column("Contractor", "user"):
        contractor_name = frappe.db.get_value("Contractor", {"user": frappe.session.user, "active_status": 1}, "name")
    if not contractor_name and is_internal:
        return contractor or frappe.db.get_value("Contractor", {"active_status": 1}, "name")

    if not contractor_name:
        frappe.throw("Your user account is not linked to an active Partner Agency.", frappe.PermissionError)

    return contractor_name


@frappe.whitelist()
def get_unpaid_commission_summary(contractor=None):
    """
    Returns a lightweight summary (count + totals) for a partner agency.
    Multi-tenant isolated for Foreign Agency users.
    """
    contractor = _resolve_session_contractor(contractor)
    contractor_doc = _get_contractor(contractor)
    default_rate = flt(contractor_doc.default_commission_amount) or 0.0
    currency = contractor_doc.default_commission_currency or "SAR"
    cols = _commission_columns_present()
    has_status = cols["commission_status"]

    status_filter = (
        "(app.commission_status IS NULL OR app.commission_status = '' OR app.commission_status = 'Unpaid')"
        if has_status else "1=1"
    )

    sql = f"""
        SELECT COUNT(DISTINCT app.name) AS total_count
        FROM `tabApplicant` app
        LEFT JOIN `tabApplicant Dossier` dos ON dos.applicant = app.name
        LEFT JOIN `tabDSR` dsr ON dsr.applicant_dossier = dos.name
        LEFT JOIN `tabDSR Departure` dep ON dep.dsr = dsr.name
        WHERE
            (app.locked_contractor = %(c)s OR dos.contractor_name = %(c)s OR dsr.contractor_name = %(c)s)
            AND (app.applicant_state = 'Departed' OR dep.status = 'Departed' OR dsr.departure_status = 'Departed')
            AND {status_filter}
    """
    result = frappe.db.sql(sql, {"c": contractor}, as_dict=True)
    total_count = (result[0].total_count if result else 0) or 0

    return {
        "summary": {
            "contractor": contractor,
            "company_name": contractor_doc.company_name,
            "country": contractor_doc.country or "—",
            "contact_person": contractor_doc.contact_person or "—",
            "phone": contractor_doc.phone or contractor_doc.whatsapp or "—",
            "default_rate": default_rate,
            "currency": currency,
            "total_count": total_count,
            "total_amount": total_count * default_rate,
            "generated_on": formatdate(today(), "dd-MMM-yyyy"),
        }
    }


@frappe.whitelist()
def get_unpaid_commission_candidates_list(contractor=None, limit=30, from_date=None, to_date=None):
    """
    Returns the full candidate list with summary for the Agency Commission view.
    Multi-tenant isolated for Foreign Agency users.
    """
    contractor = _resolve_session_contractor(contractor)
    summary, candidates = get_unpaid_commission_data(
        contractor, limit=limit, from_date=from_date, to_date=to_date
    )
    return {"summary": summary, "candidates": candidates}


@frappe.whitelist()
def export_unpaid_commission_report(contractor=None, export_format="excel", limit=30, from_date=None, to_date=None, applicant_ids=None):
    """
    Streams an Excel (.xlsx) or PDF commission billing statement for download.
    Multi-tenant isolated for Foreign Agency users.
    """
    frappe.only_for(["System Manager", "Administrator", "LMS Employee", "Accounts Manager", "Foreign Agency"])

    contractor = _resolve_session_contractor(contractor)

    summary, candidates = get_unpaid_commission_data(
        contractor=contractor,
        limit=limit,
        from_date=from_date,
        to_date=to_date,
        applicant_ids=applicant_ids,
    )

    if not candidates:
        frappe.throw(frappe._("No unpaid departed candidates found for the selected agency and filters."))

    safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in contractor)
    date_tag = today().replace("-", "")
    limit_int, _ = _normalize_limit(limit)
    batch_tag = f"Last{limit_int}" if limit_int else "All"

    fmt = str(export_format).lower()
    if fmt in ("pdf", "invoice"):
        content, mime, ext = build_commission_pdf(summary, candidates)
        filename = f"Commission_{safe_name}_{batch_tag}_{date_tag}.pdf"
    else:
        content, mime, ext = build_commission_excel(summary, candidates)
        filename = f"Commission_{safe_name}_{batch_tag}_{date_tag}.{ext}"

    frappe.response["type"] = "binary"
    frappe.response["filename"] = filename
    frappe.response["filecontent"] = content


@frappe.whitelist()
def mark_commissions_as_paid(contractor, applicant_ids=None, reference=None, payment_date=None, limit=30):
    """
    Marks a batch of departed candidates as commission-settled.
    Posts an Income entry in the applicant financial ledger.
    Requires System Manager or Accounts Manager role.
    """
    frappe.only_for(["System Manager", "Accounts Manager"])

    ids = _parse_applicant_ids(applicant_ids)
    if not ids:
        _, candidates = get_unpaid_commission_data(contractor, limit=limit)
        ids = [c["name"] for c in candidates]

    if not ids:
        return {
            "status": "warning",
            "message": frappe._("No eligible unpaid candidates found for this agency."),
        }

    if not reference:
        frappe.throw(frappe._("A payment reference (bank transfer / receipt number) is required."))

    contractor_doc = _get_contractor(contractor)
    default_rate = flt(contractor_doc.default_commission_amount) or 0.0
    currency = contractor_doc.default_commission_currency or "SAR"
    pay_date = payment_date or today()
    ref_str = reference.strip()

    # Resolve which columns exist once, outside the loop
    cols = _commission_columns_present()

    updated = 0
    total_paid = 0.0
    errors = []

    for app_id in ids:
        try:
            if not frappe.db.exists("Applicant", app_id):
                continue

            app = frappe.get_doc("Applicant", app_id)
            rate = flt(app.get("commission_amount")) or default_rate

            if cols["commission_status"]:
                app.commission_status = "Paid"
            if cols["commission_amount"]:
                app.commission_amount = rate
            if cols["commission_paid_date"]:
                app.commission_paid_date = pay_date
            if cols["commission_batch_ref"]:
                app.commission_batch_ref = ref_str

            # Idempotent income log: only append if this reference isn't already posted
            already_logged = any(
                (row.get("transaction_type") == "Income" and ref_str in (row.get("description") or ""))
                for row in (app.income_expense_logs or [])
            )
            if not already_logged:
                app.append("income_expense_logs", {
                    "transaction_type": "Income",
                    "amount": rate,
                    "date": pay_date,
                    "description": f"Agency commission — {contractor} (Ref: {ref_str})",
                    "source_doctype": "Contractor",
                    "source_name": contractor,
                })

            app.save(ignore_permissions=True)
            updated += 1
            total_paid += rate

        except Exception as exc:
            frappe.log_error(
                title=f"Commission settlement failed for {app_id}",
                message=frappe.get_traceback(),
            )
            errors.append(app_id)

    frappe.db.commit()

    msg = frappe._(
        "{0} candidates settled — {1:,.2f} {2} posted. Reference: {3}"
    ).format(updated, total_paid, currency, ref_str)

    if errors:
        msg += frappe._(" ({0} records had errors and were skipped — check Error Log.)").format(len(errors))

    return {
        "status": "success" if not errors else "partial",
        "updated_count": updated,
        "skipped_count": len(errors),
        "total_amount": total_paid,
        "currency": currency,
        "reference": ref_str,
        "message": msg,
    }
